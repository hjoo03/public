import requests, ray, time, json
from openpyxl import Workbook
from datetime import datetime, timedelta

# Latest Modified Date : 2022/03/21
# Mafia42 lastdiscussion fetch -> save as *.xlsx file
# v4.6 Patch Notes
# 1. removed date, time to speed up fetch()
# 2. Added if/else sequence before counting estimated time to avoid ZeroDivisionError

# 5.0 update -> remove duplicates to make user dictionary
# lastdiscussionv4.py => fetchuser.py

version = "5.0d"

start = 870000
end = 879604

data = ["nickname", "user_id", "lastlogin_time"]
date = str(datetime.now())[5:7] + str(datetime.now())[8:10]
write_wb = Workbook()
write_ws = write_wb.create_sheet(f"userdata_{date}")
write_ws.append(["nickname", "id", "lastlogin_date", "lastlogin_time"])
splitlist = []
splitlist2 = []
userlist = []
f = open("userlist.txt", 'r')
ul = f.readlines()
for item in ul:
    userlist.append(item[:-1])
userlist = userlist[90000:]
f.close()

if end - start <= 20000:
    splitlist.append((start, end))

else:
    i = start
    while True:
        splitlist.append((i, i + 19999))

        if end - i < 40000:
            splitlist.append((i + 20000, end))
            break

        i += 20000

if len(userlist) <= 20000:
    splitlist2.append((0, len(userlist)-1))
else:
    i = 0
    while True:
        splitlist2.append((i, i + 19999))

        if len(userlist)-1 - i < 40000:
            splitlist2.append((i + 20000, len(userlist)-1))
            break
        i += 20000
ray.init(num_cpus=8)


@ray.remote
def fetch_article(article_id):
    try:
        reqraw = requests.get(f"https://mafia42.com/api/show-lastDiscussion/{article_id}")

    except OSError:
        print(f"[ERROR][{str(datetime.now())[:len(str(datetime.now())) - 7]}] OSError: letting server rest")
        time.sleep(15)
        reqraw = requests.get(f"https://mafia42.com/api/show-lastDiscussion/{article_id}")

    try:
        res = json.loads(reqraw.text)
        res = res['boardData']

    except KeyError:
        return
    return [res["nickname"], res["user_id"], "null", "null"]


@ray.remote
def fetch_user(user_id):
    payload = {"id": user_id}
    try:
        r = requests.post(f"https://mafia42.com/api/user/user-info", data=payload)
    except OSError:
        print(f"[ERROR][{str(datetime.now())[:len(str(datetime.now())) - 7]}] OSError: letting server rest")
        time.sleep(15)
        r = requests.post(f"https://mafia42.com/api/user/user-info", data=payload)

    try:
        if r.status_code == 200:
            res = json.loads(r.text)
            res = res['userData']
        else:
            return
    except json.decoder.JSONDecodeError:
        return
    return [res["NICKNAME"], res["ID"], res["lastlogin_time"][:10], res["lastlogin_time"][11:19]]


def analyze(input_data, col):  # name, user_id, date, time
    global check
    if check:
        if str(input_data[1]) in userlist:
            return 0
    write_ws.cell(col, 1, input_data[0])
    write_ws.cell(col, 2, input_data[1])
    write_ws.cell(col, 3, input_data[2])
    write_ws.cell(col, 4, input_data[3])
    return 1


def main(x, y):
    global cnt, column
    starttime = datetime.now()
    results = [fetch_article.remote(x) for x in range(x, y)]
    while len(results):
        done, results = ray.wait(results)
        cnt += 1

        if cnt % 5000 == 0:
            print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Fetched {cnt}/{end - start + 1 + len(userlist)}")

        r = ray.get(done[0])
        if r:
            column += 1 if analyze(r, column) else 0

    cnt += 1
    endtime = datetime.now()
    ups = round((y - x + 1) / (endtime - starttime).seconds, 1) if endtime - starttime != 0 else 0
    estimatedmin = round(((end+len(userlist) - start) - cnt)/ups/60, 1) if ups != 0 else 0
    ett = str(datetime.now() + timedelta(minutes=estimatedmin))
    print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Fetched {cnt}/{end-start+1+len(userlist)} ({ups}users/sec, {estimatedmin}mins remaining)")
    print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Estimated endtime: {ett[:len(ett)-7]}")
    ray.shutdown()


def main2(x, y):
    global cnt, column
    starttime = datetime.now()
    results = [fetch_user.remote(userlist[x]) for x in range(x, y)]

    while len(results):
        done, results = ray.wait(results)
        cnt += 1

        if cnt % 5000 == 0:
            print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Fetched {cnt}/{end - start + 1 +len(userlist)}")

        r = ray.get(done[0])

        if r:
            column += 1 if analyze(r, column) else 0

    cnt += 1
    endtime = datetime.now()
    ups = round((y - x + 1) / (endtime - starttime).seconds, 1) if endtime - starttime != 0 else 0
    estimatedmin = round(((len(userlist)+end - start) - cnt)/ups/60, 1) if ups != 0 else 0
    ett = str(datetime.now() + timedelta(minutes=estimatedmin))
    print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Fetched {cnt}/{len(userlist)+end-start+1} ({ups}users/sec, {estimatedmin}mins remaining)")
    print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Estimated endtime: {ett[:len(ett)-7]}")
    ray.shutdown()


cnt = 1
fcnt = 1
column = 1
print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Splited(new): {splitlist2}")
print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Splited: {splitlist}")
check = False
for (s, e) in splitlist2:
    if __name__ == "__main__":
        main2(s, e)

    if fcnt % 5 == 0:
        write_wb.save(f"C:/Users/Joo/Documents/joo/main/userdb_{start}_{cnt}.xlsx")
        print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] {cnt}/{end - start + 1} Saved ({round(cnt / (len(userlist)+end - start) * 100, 1)}%)")
    fcnt += 1

check = True
write_wb.save(f"C:/Users/Joo/Documents/joo/main/userdb_{date}_olduser.xlsx")
print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] Initiating main()")
for (s, e) in splitlist:
    if __name__ == "__main__":
        main(s, e)

    if fcnt % 5 == 0:
        write_wb.save(f"C:/Users/Joo/Documents/joo/main/userdb_{start}_{cnt}.xlsx")
        print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] {cnt}/{len(userlist)+end - start + 1} Saved ({round(cnt / (len(userlist)+end - start) * 100, 1)}%)")
    fcnt += 1


write_wb.save(f"C:/Users/Joo/Documents/joo/main/userdb_{cnt - 1}.xlsx")
print(f"[{str(datetime.now())[:len(str(datetime.now())) - 7]}] {cnt - 1}/{len(userlist)+end - start + 1} Saved ({round(cnt / (len(userlist)+end - start) * 100, 1)}%)")
