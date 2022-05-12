import requests, sys, datetime, selenium, time
from bs4 import BeautifulSoup as Bs
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from PyQt5 import uic
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from openpyxl import Workbook
form_class = uic.loadUiType("shopping_main.ui")[0]


class MainWindow(QMainWindow, form_class):

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.validity = False
        self.wd_options = webdriver.ChromeOptions()
        self.wd_options.add_argument("headless")
        self.wd_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), options=self.wd_options)
        self.prod_list = []
        self.col = 2
        self.store = ""

        self.urlinput.textChanged.connect(self.store_input)
        self.searchbtn.clicked.connect(self.fetch)

        self.write_wb = Workbook()
        self.write_ws = None
        self.already = False

    def fetch(self):
        if self.already:
            self.warning_error03()
            return

        if not self.store:
            self.warning_error02()
            return

        sheet_title = f'{self.store}_{str(datetime.datetime.now())[:len(str(datetime.datetime.now())) - 7][2:10].replace("-", "")}'
        self.write_ws = self.write_wb.create_sheet(sheet_title)
        self.write_ws.append(["상품번호", "상품명", "가격", "배송비", "옵션"])
        url = "https://smartstore.naver.com/" + self.store
        res = requests.get(url)

        if res.status_code != 200:
            self.warning_error01(res.status_code)
            return

        html = res.text
        soup = Bs(html, "html.parser")

        cnt = 1
        while True:
            try:
                product_raw = soup.select_one(f"#pc-wholeProductWidget > div > div > div > ul > li:nth-child({cnt}) > a").get("href")
                product_no = product_raw.split('/')[3]
                self.prod_list.append([product_no])
                cnt += 1

            except AttributeError:
                break

        self.output("Appended {0} products to queue".format(len(self.prod_list)))
        cnt = 1
        for product in self.prod_list:
            url_prod = f"https://smartstore.naver.com/{self.store}/products/" + str(product[0])
            res_prod = requests.get(url_prod)

            if res.status_code != 200:
                self.warning_error01(res_prod.status_code)
                return

            html_prod = res_prod.text
            soup_prod = Bs(html_prod, "html.parser")
            prod_title = soup_prod.select_one("#content > div > div._2-I30XS1lA > div._2QCa6wHHPy > fieldset > div._1ziwSSdAv8 > div.CxNYUPvHfB > h3").text
            self.driver.get(url_prod)
            time.sleep(0.5)
            options = {}
            price = self.driver.find_element(By.CSS_SELECTOR,
                                             "#content > div > div._2-I30XS1lA > div._2QCa6wHHPy > fieldset > div._1ziwSSdAv8 > div.WrkQhIlUY0 > div > strong > span._1LY7DqCnwR").text
            delivery = self.driver.find_element(By.CSS_SELECTOR,
                                                "#content > div > div._2-I30XS1lA > div._2QCa6wHHPy > fieldset > div._1rGSKv6aq_ > div > span:nth-child(2)").text

            os_cnt = 1
            validity = True
            while validity:  # 옵션 추가
                try:
                    self.driver.find_element(By.CSS_SELECTOR,
                                             f"#content > div > div._2-I30XS1lA > div._2QCa6wHHPy > fieldset > div.bd_2dy3Y > div:nth-child({os_cnt}) > a").click()  # 옵션 펼치기
                    time.sleep(0.3)

                    category = self.driver.find_element(By.CSS_SELECTOR,
                                                        f"#content > div > div._2-I30XS1lA > div._2QCa6wHHPy > fieldset > div.bd_2dy3Y > div:nth-child({os_cnt}) > a").text

                    options[category] = []

                    o_cnt = 1
                    validity_2 = True
                    while validity_2:
                        op_text = self.driver.find_element(By.CSS_SELECTOR,
                                                           f"#content > div > div._2-I30XS1lA > div._2QCa6wHHPy > fieldset > div.bd_2dy3Y > div:nth-child({os_cnt}) > ul > li:nth-child({o_cnt}) > a").text  # 옵션 내용
                        if not op_text:
                            validity_2 = False
                        else:
                            options[category].append(op_text)
                        o_cnt += 1

                    os_cnt += 1
                except selenium.common.exceptions.NoSuchElementException:
                    self.prod_list[cnt - 1].append([prod_title, price, delivery, options])
                    cnt += 1
                    validity = False

        self.excel(self.prod_list)
        self.output("Saved to D:/test.xlsx")
        self.already = True

    def store_input(self):
        self.store = self.urlinput.text()

    def excel(self, data):
        for [n, [t, p, d, o]] in Win.prod_list:
            self.write_ws.cell(self.col, 1, n)
            self.write_ws.cell(self.col, 2, t)
            self.write_ws.cell(self.col, 3, p)
            self.write_ws.cell(self.col, 4, d)

            for category in o:
                data = ""
                data += category + ": "
                for item in o[category]:
                    data += item + ", "
                data = data[:-2] + "\n"

            self.write_ws.cell(self.col, 5, data)

            self.col += 1
        self.write_ws.cell(self.col, 1, "Updated:")
        self.write_ws.cell(self.col, 2, str(datetime.datetime.now())[:len(str(datetime.datetime.now())) - 7])
        self.write_wb.save(f"D:/test.xlsx")

    def warning_error01(self, args):
        QMessageBox.warning(self, "Request Error", f"Response Code: {args}")

    def warning_error02(self):
        QMessageBox.warning(self, "InputError", f"No Input!")

    def warning_error03(self):
        QMessageBox.warning(self, "Error", f"다중검색은 미구현 기능입니다.")

    def output(self, msg):
        self.console.append(msg)

    def keyPressEvent(self, k):
        if k.key() == Qt.Key_Return:
            self.fetch()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    Win = MainWindow()
    Win.show()
    app.exec_()
