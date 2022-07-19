# -*- coding: utf-8 -*-
#
# Resolution: 1920 × 1080 (FHD)
# Chrome Only


import os, sys, time, subprocess, openpyxl, pyautogui, shutil, datetime, logging
from logging.handlers import RotatingFileHandler
import selenium.common.exceptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from PyQt5.QtWidgets import *
from selenium import webdriver
from subprocess import CREATE_NO_WINDOW  # Only available in Windows
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from PyQt5.QtGui import QIntValidator
from openpyxl.drawing.image import Image
from urllib import parse
from PyQt5.QtCore import QThread, QObject, pyqtSignal
from form import Ui_SettingsWindow


class Logger:
    def __init__(self):
        if not os.path.isdir(os.getcwd() + "\\log\\"):
            os.makedirs(os.getcwd() + "\\log")

        self.filename = os.getcwd() + '\\log\\' + datetime.datetime.now().strftime('%y%m%d_%H%M') + '.log'
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.INFO)
        self.formatter = logging.Formatter(u'%(asctime)s [%(levelname)s] File "%(filename)s", line %(lineno)d, in %(funcName)s: "%(message)s"')
        self.file_handler = logging.FileHandler(self.filename)

        log_max_size = 10 * 1024 * 1024  # 10MB
        log_file_count = 20
        rotatingFileHandler = logging.handlers.RotatingFileHandler(
            filename=self.filename,
            maxBytes=log_max_size,
            backupCount=log_file_count
        )
        rotatingFileHandler.setFormatter(self.formatter)
        self.logger.addHandler(rotatingFileHandler)

    def debug(self, msg):
        self.logger.debug(msg)

    def info(self, msg):
        self.logger.info(msg)

    def error(self, msg):
        self.logger.error(msg)


class WebDriver:
    def __init__(self):
        subprocess.Popen(r'C:\Program Files\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\chrometemp"')
        option = Options()
        option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        option.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36")

        if getattr(sys, 'frozen', False):
            chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
            self.chrome_service = ChromeService(chromedriver_path)
        else:
            self.chrome_service = ChromeService("chromedriver.exe")
        self.chrome_service.creationflags = CREATE_NO_WINDOW
        self.driver = webdriver.Chrome(service=self.chrome_service, options=option)
        self.action = ActionChains(self.driver)
        self.driver.maximize_window()
        log.info("Webdriver Initiated")

        """time.sleep(2)
        self.driver.get("https://world.taobao.com/wow/z/oversea/SEO-SEM/ovs-pc-login")
        wait = WebDriverWait(self.driver, 5)
        _ = wait.until(ec.visibility_of_element_located((By.XPATH, "//*[@id=\"login-form\"]/div[4]/button")))
        time.sleep(5)
        log.info("Login Success")"""

        self.search_button_path = "/html/body/div/div/div[1]/div[1]/div[1]"
        self.try_again_button_path = "//*[@id=\"ap-sbi-taobao-result\"]/div/div[2]/div/div[1]/div[2]/div"
        self.login_button_path = "//*[@id=\"ap-sbi-taobao-result\"]/div/div[2]/div/div[2]/div[2]/div"
        self.verify_button_path = "//*[@id=\"ap-sbi-taobao-result\"]/div/div[2]/div/div[3]/div[2]/div"

    def main(self, cell, limit, sheet):
        """
        get request image file at the driver
        then searches it on taobao
        :param cell: str, target cell
        :param limit: int, limit of items per image
        :param sheet: Workbook.Sheet(), for preventing I/O on closed file
        :return data: json, {response_code: CODE, data: list of [title, price, sales]}
        """
        img_link = sheet[cell].value
        self.driver.get(img_link)
        return self.search(limit)

    def search(self, limit):
        data = self.taobao_extension(limit)
        if not data:  # Unsuccessful Search
            SW.status.setStyleSheet("Color : red")
            SW.status.setText("Skipping...")
            time.sleep(3)
            data = self.analyze(limit)
            if data:
                SW.status.setStyleSheet("Color : green")
                SW.status.setText("Normal")
                return {"response_code": 916, "data": data}

            if data:
                return {"response_code": 916, "data": data}
            else:
                return {"response_code": 1}

        return {"response_code": 916, "data": data}

    def taobao_extension(self, limit):
        self.action.move_to_element(self.driver.find_element(By.XPATH, f"/html/body/img")).perform()
        time.sleep(0.05)
        try:
            self.driver.find_element(By.XPATH, self.search_button_path).click()
        except selenium.common.exceptions.ElementNotInteractableException:
            print("DO NOT Move Mouse Pointer!! Retrying...")
            time.sleep(0.2)
            self.action.move_to_element(self.driver.find_element(By.XPATH, f"/html/body/img")).perform()
            time.sleep(0.05)
            self.driver.find_element(By.XPATH, self.search_button_path).click()
        try:
            wait = WebDriverWait(self.driver, 15)
            _ = wait.until(ec.visibility_of_element_located((By.CSS_SELECTOR, self.path_analyzer(1)[0])))
            time.sleep(0.5)
        except selenium.common.exceptions.TimeoutException:
            return

        return self.analyze(limit)

    def analyze(self, limit):
        data = []
        for i in range(1, limit+1):  # limits amount of items to fetch (Default: 30)
            end_of_list = False
            paths = self.path_analyzer(i)
            output = []  # title, price, sales
            cnt = 0
            for path in paths:
                try:
                    res = self.driver.find_element(By.CSS_SELECTOR, path).text
                    if cnt == 1:
                        res = float(res[1:].replace(',', ''))
                    elif cnt == 2:
                        try:
                            res = int(res[:-3])
                        except ValueError:
                            res = 0

                    output.append(res)
                except selenium.common.exceptions.NoSuchElementException:
                    end_of_list = True
                    break
                cnt += 1
            if output:
                data.append(output)
            if end_of_list:
                break

        return data

    def add_link(self, data):
        """
        gets link and analyzes data
        then saves image at temporary directory
        :param data: tuple, (row, [item1, item2], [price1, sales1], [price2, sales2], extra_data)
        # extra_data: [index, title, sales, price]
        :return analyzed_data: tuple, (row, [link1, "2, sales1, "2, price1, "2], extra_data)
        """
        extra_data = []
        for [index, title, sales, price] in data[4]:
            self.driver.find_element(By.CSS_SELECTOR, self.path_analyzer(index+1)[0]).click()
            time.sleep(1)
            self.driver.switch_to.window(self.driver.window_handles[1])
            link = self.parse_url(self.driver.current_url)
            self.close_tab_from_back()
            extra_data.append([link, title, sales, price])

        links = []
        for i in range(2):
            self.driver.find_element(By.CSS_SELECTOR, self.path_analyzer(data[1][i]+1)[0]).click()
            time.sleep(1)
            self.driver.switch_to.window(self.driver.window_handles[1])
            time.sleep(0.2)
            links.append(self.parse_url(self.driver.current_url))
            self.close_tab_from_back()
            img_path = f"#ap-sbi-taobao-result > div > div.ap-list.ap-list--gallery > div > div.simplebar-wrapper > div.simplebar-mask > div > div > div > div > div:nth-child({data[1][i]+1}) > div > div.ap-is-link.ap-product-image > img"
            # imgUrl = self.driver.find_element(By.CSS_SELECTOR, img_path).get_attribute("src")
            self.download_img(img_path, os.getcwd() + rf"\temp\img\{data[0]}_{i}.jpg")

        return data[0], [links[0], links[1], data[2][1], data[3][1], data[2][0], data[3][0]], extra_data

    def close_extension(self):
        close_button = f"/html/body/div/div/div[1]/div[1]/div/div[3]/div/div[3]/div[1]"
        self.driver.find_element(By.XPATH, close_button).click()
        time.sleep(0.5)

    def download_img(self, css_path, file_path):
        with open(file_path, "wb") as file:
            img = self.driver.find_element(By.CSS_SELECTOR, css_path)
            file.write(img.screenshot_as_png)

    def close_driver(self):
        try:
            self.close_tab_from_front(len(self.driver.window_handles))
            self.driver.quit()
        except selenium.common.exceptions.InvalidSessionIdException:
            pass

    @staticmethod
    def parse_url(url):
        if "ali_redirect.html?url" in url:
            return parse.unquote(url[url.index("url=")+4:url.index("%26ns")], encoding="utf-8")
        elif "member/login.html" in url:
            return url[url.index("redirect=")+9:url.index("&ns")]
        else:
            return url

    @staticmethod
    def path_analyzer(item_no):
        base_path = f"#ap-sbi-taobao-result > div > div.ap-list.ap-list--gallery > div > div.simplebar-wrapper > div.simplebar-mask > div > div > div > div > div:nth-child({item_no}) > div > "
        price = base_path + "div.ap-tb-deal-info > div.ap-product-price"
        sales = base_path + "div.ap-tb-deal-info > div.ap-tb-sales"
        title = base_path + "div.ap-is-link.ap-product-title"
        return title, price, sales

    def close_tab_from_front(self, count=1):
        try:
            tabs = self.driver.window_handles
        except selenium.common.exceptions.InvalidSessionIdException:
            log.error("Invalid Session ID")
            return
        for i in range(count):
            self.driver.switch_to.window(tabs[0])
            self.driver.close()
            try:
                tabs = self.driver.window_handles
            except selenium.common.exceptions.InvalidSessionIdException:
                log.error("Invalid Session ID")
                return
        try:
            self.driver.switch_to.window(tabs[0])
        except selenium.common.exceptions.InvalidSessionIdException:
            log.error("Invalid Session ID")

    def close_tab_from_back(self, count=1):
        tabs = self.driver.window_handles
        for i in range(count):
            self.driver.switch_to.window(tabs[len(tabs)-1])
            self.driver.close()
            tabs = self.driver.window_handles
        self.driver.switch_to.window(tabs[0])


class Sheet:
    def __init__(self):
        if not os.path.isdir(os.getcwd() + "\\temp\\img\\"):
            os.makedirs(os.getcwd() + "\\temp\\img")

        self.doc = None
        self.sheet = None
        self.ws = None
        self.extra = 0
        self.this_sheet_items = 0
        self.items = 0

    def setup_read_sheet(self, file_directory, sheet_name):
        self.doc = openpyxl.load_workbook(file_directory)
        self.sheet = self.doc[sheet_name]

    def setup_target_sheet(self):
        self.doc = openpyxl.load_workbook(SW.t_file_dir + "/" + SW.filename + ".xlsx")
        self.ws = self.doc.active

    def write(self, data):
        """
        writes main data in the sheet
        :param data: dict, {skips: skipped rows, data: list of (row, [link1, "2, sales1, "2, price1, "2], extra_data)}
        :return:
        """
        self.setup_target_sheet()
        extra_data = []
        for item in data["data"]:
            extra_data.append(item[2])
        self.write_extra(extra_data)
        for (row, info, _) in data["data"]:
            self.ws.add_image(Image(os.getcwd() + rf"\temp\img\{row}_0.jpg"), f"I{row}")
            self.ws.add_image(Image(os.getcwd() + rf"\temp\img\{row}_1.jpg"), f"J{row}")
            self.ws[f'K{row}'].hyperlink = info[0]
            self.ws[f'L{row}'].hyperlink = info[1]
            self.ws[f'K{row}'].value = info[0]
            self.ws[f'L{row}'].value = info[1]
            self.ws[f'K{row}'].style = info[0]
            self.ws[f'L{row}'].style = info[1]
            self.ws[f'M{row}'] = f"{info[2]}/{info[4]}"
            self.ws[f'N{row}'] = f"{info[3]}/{info[5]}"
            self.items += 1

        self.doc.save(SW.t_file_dir + "\\" + SW.filename + f"_{(self.items%100+1):02}.xlsx")
        print(f"last saved: {self.items+1}")
        SW.e_saved_label.setText(f"Extra_Saved {self.extra}")
        shutil.rmtree(os.getcwd() + "\\temp\\img\\")
        os.makedirs(os.getcwd() + "\\temp\\img")

    def write_extra(self, data):
        """
        writes extra data in the sheet
        :param data: list, list of [link, title, sales, price]
        :return:
        """
        for data_ in data:
            for d in data_:
                self.ws[f'K{SW.end+self.extra+1}'] = d[0]
                self.ws[f'M{SW.end+self.extra+1}'] = f"{d[2]}/{d[3]}"
                self.ws[f'O{SW.end+self.extra+1}'] = d[1]
                self.extra += 1

    @staticmethod
    def find_sales_high(data, min_price, min_sales) -> tuple:
        """
        analyzes data
        :param data: list, [row, list of [title, price, sales]]
        :param min_price: int
        :param min_sales: int
        :return high_sales_data: tuple, (row, [item1, item2], [price1, sales1], [price2, sales2], extra_data)
        """
        data_list = data[1]
        index = 0
        extra_data = []
        for (title, price, sales) in data_list:
            if sales >= int(SW.min_b_e):
                extra_data.append([index, title, sales, price])
            index += 1

        sales_list = []
        for d in data_list[:int(SW.a_i)]:
            if (d[1] < min_price) or (d[2] < min_sales):
                sales_list.append(-1)
            else:
                sales_list.append(d[2])
        sorted_list = sorted(sales_list, reverse=True)
        s_1 = sales_list.index(sorted_list[0])
        if len(sales_list) == 1:
            return data[0], [s_1, 0], [data[1][s_1][1], data[1][s_1][2]], [0, 0], extra_data
        if sorted_list[0] == sorted_list[1]:
            salesM = sorted_list[0]
            s_2 = sales_list[sales_list.index(salesM)+1:].index(salesM) + sales_list.index(salesM) + 1
        else:
            s_2 = sales_list.index(sorted_list[1])

        return data[0], [s_1, s_2], [data[1][s_1][1], data[1][s_1][2]], [data[1][s_2][1], data[1][s_2][2]], extra_data

    def setup_target_sheet_init(self):
        self.doc = openpyxl.load_workbook(SW.t_file_dir + "/" + SW.filename + ".xlsx")
        self.ws = self.doc.active
        self.ws['I1'] = "사진1"
        self.ws['J1'] = "사진2"
        self.ws['K1'] = "링크1"
        self.ws['L1'] = "링크2"
        self.ws['M1'] = "구매수/가격1"
        self.ws['N1'] = "구매수/가격2"
        self.ws['O1'] = "타오바오 상품명"

        self.doc.save(SW.t_file_dir + "/" + SW.filename + ".xlsx")

    @staticmethod
    def copy_file():
        shutil.copy(SW.o_file_dir, SW.t_file_dir + "/" + SW.filename + ".xlsx")


def click(x, y, rl='l'):
    if rl == 'r':
        pyautogui.rightClick(x, y)
    else:
        pyautogui.click(x, y)


def timestamp():
    return f"[{str(datetime.datetime.now())[:len(str(datetime.datetime.now())) - 7]}]"


# noinspection PyArgumentList
class SettingsWindow(QMainWindow, Ui_SettingsWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.onlyInt = QIntValidator()
        self.o_file_dir = ''
        self.t_file_dir = ''
        self.filename = "file"
        self.e_saved = 0
        self.min_p = 0
        self.min_b = 0
        self.min_b_e = 200
        self.p_i = 30
        self.a_i = 10
        self.start_row = 2
        self.delay = 200
        self.savef = 5
        self.minprice.setValidator(self.onlyInt)
        self.minbuy.setValidator(self.onlyInt)
        self.minbuy_extra.setValidator(self.onlyInt)
        self.peritem.setValidator(self.onlyInt)
        self.a_item.setValidator(self.onlyInt)
        self.startrow_lb.setValidator(self.onlyInt)
        self.delayt.setValidator(self.onlyInt)
        self.save.setValidator(self.onlyInt)
        self.minprice.textChanged.connect(self.s_m1_p)
        self.minbuy.textChanged.connect(self.s_m1_b)
        self.minbuy_extra.textChanged.connect(self.s_m_e)
        self.peritem.textChanged.connect(self.s_p_i)
        self.a_item.textChanged.connect(self.s_a_i)
        self.startrow_lb.textChanged.connect(self.s_sr)
        self.filename_lb.textChanged.connect(self.set_filename)
        self.delayt.textChanged.connect(self.sd)
        self.save.textChanged.connect(self.ss)

        self.status.setStyleSheet("Color : green")
        self.pushButton.clicked.connect(self.start)  # button triggers start()
        self.file_button.clicked.connect(self.select_file)
        self.file_button_2.clicked.connect(self.select_dir)

        self.thread, self.worker = None, None
        self.end = 0
        self.now = 0
        self.skipped = []
        self.consecutive_skips = 0

    # noinspection PyUnresolvedReferences
    def start(self):
        self.thread = QThread()
        self.worker = Worker()
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        if not self.o_file_dir:
            self.warning_error01()
            return
        if not self.t_file_dir:
            self.warning_error02()
            return

        self.worker.St.setup_read_sheet(SW.o_file_dir, "export")  # Excel file dir, Sheet name
        self.worker.St.copy_file()
        s = int(self.start_row)
        while True:
            if not self.worker.St.sheet[f'A{s}'].value and self.worker.St.sheet[f'A{s}'].value != 0:
                self.end = s - 1
                break
            s += 1
        SW.saved_label.setText(f"Saved {SW.now}/{self.end - 1}")
        self.worker.St.setup_target_sheet_init()
        self.thread.start()

    def select_file(self):
        file_filter = "Excel File (*.xlsx *.xls)"
        self.o_file_dir = QFileDialog.getOpenFileName(
            parent=self,
            caption="Select an Excel File",
            directory=os.getcwd(),
            filter=file_filter,
            initialFilter=file_filter
        )[0]
        self.file_label.setText(self.o_file_dir)
        self.file_label.repaint()
        self.filename_lb.setText(self.o_file_dir[self.o_file_dir.rindex('/')+1:self.o_file_dir.rindex('.xlsx')] + "_result")
        self.filename_lb.repaint()
        self.filename = self.o_file_dir[self.o_file_dir.rindex('/')+1:self.o_file_dir.rindex('.xlsx')] + "_result"

    def select_dir(self):
        self.t_file_dir = QFileDialog.getExistingDirectory(
            parent=self,
            caption="Select a folder"
        )
        self.file_label_2.setText(self.t_file_dir)
        self.file_label_2.repaint()

    def s_m1_p(self):
        self.min_p = self.minprice.text()

    def s_m1_b(self):
        self.min_b = self.minbuy.text()

    def s_m_e(self):
        self.min_b_e = self.minbuy_extra.text()

    def s_p_i(self):
        self.p_i = self.peritem.text()

    def s_a_i(self):
        self.a_i = self.a_item.text()

    def s_sr(self):
        self.start_row = self.startrow_lb.text()

    def sd(self):
        self.delay = self.delayt.text()

    def ss(self):
        self.savef = self.save.text()

    def set_filename(self):
        self.filename = self.filename_lb.text()

    def warning_error01(self):
        QMessageBox.warning(self, "[E01] FileError", "No File Selected!")

    def warning_error02(self):
        QMessageBox.warning(self, "[E02] FileError", "No Target Directory Selected!")

    def warning_error03(self, ts):
        QMessageBox.warning(self, "[E03] UnknownError", f"Unknown Error occurred at [{ts}]")


class Worker(QObject):
    finished = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.St = Sheet()
        self.WD = None
        self.accumulated_data = {"skips": [], "data": []}
        self.deleted = 0

    def run(self):
        cycles = 0
        for rows in range(int(SW.start_row), SW.end + 1, int(SW.savef)):
            res = self.main(rows)
            if res["response_code"] == 2:
                # noinspection PyUnresolvedReferences
                self.finished.emit()
                return

            cycles += 1

            if cycles % 4 == 0:
                time.sleep(int(SW.delay))

        os.remove(SW.t_file_dir + "/" + SW.filename + ".xlsx")
        # noinspection PyUnresolvedReferences
        self.finished.emit()

    def main(self, rows):
        self.WD = WebDriver()
        self.accumulated_data = {"skips": [], "data": []}

        for row in range(rows, rows + int(SW.savef)):
            row_ = row - self.deleted
            res = self.wd_data(row_)
            if res["response_code"] == 2:
                log.error("3 consecutive skips. return error")
                return res
            elif res["response_code"] == 3:
                self.St.ws.delete_rows(row_, 1)
                self.deleted += 1

        self.St.write(self.accumulated_data)
        log.info(f"Saved {len(self.accumulated_data['data'])} items")
        SW.saved_label.setText(f"Saved {SW.now}/{SW.end - 1}")

        self.WD.close_driver()
        del self.WD
        time.sleep(2)
        return {"response_code": 916}

    def wd_data(self, row):
        d = self.WD.main(f"B{row}", SW.p_i, self.St.sheet)
        if d["response_code"] != 916:
            SW.status.setStyleSheet("Color : green")
            SW.status.setText("Normal")
            SW.consecutive_skips += 1
            if SW.consecutive_skips == 2:
                time.sleep(180)
            if SW.consecutive_skips > 2:
                SW.startrow_lb.setText(str(int(SW.now)-3))
                SW.start_row = int(SW.now) - 3
                SW.warning_error03(timestamp())
                return {"response_code": 2}

            return {"response_code": 3}

        data = self.WD.add_link(self.St.find_sales_high([row, d["data"]], SW.min_p, SW.min_b))
        self.WD.close_extension()
        SW.now += 1
        SW.current_label.setText(f"Current {SW.now}/{SW.end - 1}")
        self.accumulated_data["data"].append(data)
        SW.consecutive_skips = 0
        return {"response_code": 916}


if __name__ == "__main__":
    app = QApplication(sys.argv)
    SW = SettingsWindow()
    log = Logger()
    SW.show()
    app.exec_()
