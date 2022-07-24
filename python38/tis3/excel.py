# -*- coding: utf-8 -*-
# tis/excel.py

import os, openpyxl, shutil, string
from openpyxl.drawing.image import Image
from openpyxl_image_loader import SheetImageLoader
from logger import Logger


class Excel:
    def __init__(self, fd):
        global log
        log = Logger(fd, "excel").logger
        if not os.path.isdir(os.getcwd() + "\\temp\\img\\"):
            os.makedirs(os.getcwd() + "\\temp\\img")
        self.worksheet, self.doc, self.ws, self.read_sheet = None, None, None, None
        self.ex_exists = False
        self.extra = 0
        self.finished_items = 0
        self.in_directory = ''
        self.out_directory = ''
        self.filename = ''
        self.index_range = []
        self.skipped_list = []
        self.tmp = os.getcwd() + "\\_temp.xlsx"
        self.tmp_h = os.getcwd() + "\\_h_temp.xlsx"
        self.res = self.out_directory + self.filename + ".xlsx"
        self.res_h = self.out_directory + self.filename + "_h.xlsx"

    def set_directory(self, in_dir, out_dir, filename):
        self.in_directory = in_dir
        self.out_directory = out_dir + '\\'
        self.filename = filename
        self.read_sheet = openpyxl.load_workbook(self.in_directory).active
        self.create_template()

    def setup_sheet(self):
        self.doc = openpyxl.load_workbook(self.tmp)
        self.ws = self.doc.active

    def setup_extra_sheet(self):
        self.doc = openpyxl.load_workbook(self.tmp_h)
        self.ws = self.doc.active

    def setup_new_sheet(self):
        self.doc = openpyxl.Workbook()
        self.ws = self.doc.active

    def write(self, data: dict, row: int):
        """
        writes main data in the sheet
        :param row: int, for splitting files
        :param data: tuple, (index, [link1, "2, sales1, price1, sales2, price2], extra_data)
        :return index: int
        """
        index = data[0]
        extra_data = [data[2]]
        self.write_extra(extra_data, index)

        self.setup_sheet()
        info = data[1]
        img0 = Image(os.getcwd() + rf"\temp\img\{index}_0.jpg")
        img0.width, img0.height = 132, 132
        img1 = Image(os.getcwd() + rf"\temp\img\{index}_1.jpg")
        img1.width, img1.height = 132, 132
        self.ws.add_image(img0, f"I{row}")
        self.ws.add_image(img1, f"J{row}")
        self.ws[f'K{row}'].hyperlink = info[0]
        self.ws[f'L{row}'].hyperlink = info[1]
        self.ws[f'K{row}'].value = info[0]
        self.ws[f'L{row}'].value = info[1]
        self.ws[f'K{row}'].style = "Hyperlink"
        self.ws[f'L{row}'].style = "Hyperlink"
        self.ws[f'M{row}'] = f"{info[2]}/{info[3]}"
        self.ws[f'N{row}'] = f"{info[4]}/{info[5]}"
        self.doc.save(self.tmp)
        self.finished_items += 1

        log.info(f"Wrote item_{index} at row {row}; total: {self.finished_items}")
        shutil.rmtree(os.getcwd() + "\\temp\\img\\")
        os.makedirs(os.getcwd() + "\\temp\\img")

        return index

    def write_extra(self, data, index):
        """
        writes extra data in the sheet
        :param index: int
        :param data: list, [list of [link, title, sales, price]]
        :return:
        """
        if not self.ex_exists:
            self.create_extra_sheet()
        self.setup_extra_sheet()

        for data_ in data:
            for d in data_:
                self.ws[f'A{self.extra + 2}'] = self.extra
                self.ws[f'C{self.extra + 2}'].hyperlink = d[0]
                self.ws[f'C{self.extra + 2}'].value = d[0]
                self.ws[f'C{self.extra + 2}'].style = "Hyperlink"
                self.ws[f'B{self.extra + 2}'] = d[1]
                self.ws[f'D{self.extra + 2}'] = d[2]
                self.ws[f'E{self.extra + 2}'] = d[3]
                self.extra += 1
                log.info(f"Added item_{index} to high-sales; total: {self.extra}")
        self.doc.save(self.tmp_h)

    def delete_blanks(self, start, end):
        self.setup_sheet()

        c = 0
        for row in range(start, end + 1):
            r = row + c
            if not self.ws[f'K{r}'].value:
                self.delete_images(r, 1, self.ws)
                self.ws.delete_rows(r, 1)
                c -= 1
        self.doc.save(self.tmp)

    def create_extra_sheet(self):
        self.setup_new_sheet()
        self.ws.append(["Index", "타오바오 상품명", "링크", "구매수", "가격"])
        self.doc.save(self.tmp_h)
        self.ex_exists = True

    def create_template(self):
        self.doc = openpyxl.load_workbook(self.in_directory)
        self.ws = self.doc.active
        self.ws['I1'] = "사진1"
        self.ws['J1'] = "사진2"
        self.ws['K1'] = "링크1"
        self.ws['L1'] = "링크2"
        self.ws['M1'] = "구매수/가격1"
        self.ws['N1'] = "구매수/가격2"
        self.ws['O1'] = "타오바오 상품명"
        self.doc.save(self.tmp)

    @staticmethod
    def extract_image(index, cell, sheet):  # openpyxl_image_loader.SheetImageLoader
        image_loader = SheetImageLoader(sheet)
        image = image_loader.get(cell)
        rgb_image = image.convert("RGB")
        rgb_image.save(os.getcwd() + rf"\temp\img_{index}.jpg")
        del image, rgb_image, image_loader
        return os.getcwd() + rf"\temp\img_{index}.jpg"

    def index_to_row(self, index: int) -> int:
        for row in self.index_range:
            if index == int(self.read_sheet[f'A{row}'].value):
                return row

    def row_to_index(self, row: int) -> int:
        return int(self.read_sheet[f'A{row}'].value)

    @staticmethod
    def delete_images(start_row, amount, sheet):
        sheet_images = sheet._images[:]

        _images = {}
        for image in sheet_images:
            row = image.anchor._from.row + 1
            col = string.ascii_uppercase[image.anchor._from.col]
            cell = f'{col}{row}'

            _images[cell] = image

        if amount == 0:
            row = start_row
            while True:
                try:
                    sheet._images.remove(_images[f'B{row}'])
                    row += 1
                except KeyError:
                    break
        else:
            for row in range(start_row, start_row + amount):
                try:
                    sheet._images.remove(_images[f'B{row}'])
                except KeyError:
                    break

    """
    def split_file(self, splits, start_row, total_items):
        split_list = []
        j = total_items // splits
        for i in range(0, j + 1):
            if i == j:
                split_list.append((start_row + splits * i, start_row + total_items))
            else:
                split_list.append((start_row + splits * i, start_row + splits * (i + 1)))

        for (cnt, (s, e)) in enumerate(split_list, start=1):
            wb = openpyxl.load_workbook(self.out_directory + self.filename + f"_{self.current_file:02}.xlsx")
            ws = wb.active

            del_rows = [(start_row, s), (e, start_row + total_items)]
            a = del_rows[0][1] - del_rows[0][0]
            b = del_rows[1][1] - del_rows[1][0]
            if a != 0:
                ws.delete_rows(del_rows[0][0], a)
            if b != 0:
                ws.delete_rows(del_rows[1][0], b)
            self.delete_images(del_rows[0][0], del_rows[0][1] - del_rows[0][0], ws)
            self.delete_images(del_rows[1][0], 0, ws)

            images = [self.extract_image(index, f"B{row}", self.read_sheet) for (index, row) in enumerate(range(s, e), start=2)]
            for (row, read_row) in enumerate(range(s, e), start=2):
                img = Image(images[row-2])
                img.height = 132
                img.width = 132
                ws.add_image(img, f"B{row}")
                for column in ['A', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws[column + str(row)] = self.read_sheet[column + str(read_row)].value
            wb.save(self.out_directory + self.filename + f"_{cnt:02}.xlsx")
            log.info(f"Split File {cnt}/{len(split_list)}")
            shutil.rmtree(os.getcwd() + "\\temp")
            os.makedirs(os.getcwd() + "\\temp\\img")

            self.setup_now_sheet()
    """
