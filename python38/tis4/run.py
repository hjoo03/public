# -*- coding: utf-8 -*-
# tis4/run.py

import os
import sys
import time
import datetime
import shutil
import threading

import openpyxl
import pyautogui
import pyperclip

from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox, QApplication
from PyQt5.QtCore import QThread
from PyQt5.QtGui import QIntValidator, QTextCursor
from openpyxl.drawing.image import Image

from logger import Logger
from form import Ui_Window
import manipulate_exif

# BASE_DIR = os.path.dirname(os.path.abspath(__file__)) + '\\'
BASE_DIR = os.path.dirname(os.path.abspath(__file__)) + "\\images\\"

class Error(OSError):
    pass

class Macro:
    main_window = (950, 60)
    btn_camera = (342, 135)
    btn_search = (227, 670)
    btn_gallery = (42, 735)
    btn_gallery_lower = (42, 840)
    item_1 = (70, 220)
    result_1 = (110, 350, 1)
    result_2 = (340, 350, 2)
    btn_share = (332, 90)
    btn_copy = (42, 515)
    btn_escape = (350, 930)
    btn_escapeItemDetail = (25, 90)

    def __init__(self):
        pass

    def move_window(self):
        pyautogui.moveTo(self.main_window[0], self.main_window[1])
        pyautogui.dragTo(210, 20, duration=0.25)

    @staticmethod
    def click(coordinate):
        pyautogui.moveTo(coordinate[0], coordinate[1], 0.25)
        pyautogui.doubleClick()
        # pyautogui.click(coordinate[0], coordinate[1])

    @staticmethod
    def single_click(coordinate):
        pyautogui.moveTo(coordinate[0], coordinate[1], 0.25)
        pyautogui.click()

    @staticmethod
    def scroll(constant: int = 0):
        pyautogui.moveTo(200, 626)
        pyautogui.dragTo(200, 500 + constant, duration=1)

    @staticmethod
    def before_locate():
        pyautogui.moveTo(2, 2, 0.1)
        time.sleep(0.1)

    @staticmethod
    def drag():  # TODO: needs to be fixed
        pyautogui.moveTo(90, 620, 0.1)
        time.sleep(0.5)
        pyautogui.dragTo(410, 645, 2.5, pyautogui.easeInOutQuad)

class Excel:
    def setup_sheet(self):
        self.doc = openpyxl.load_workbook(self.tmp)
        self.ws = self.doc.active

    def write(self, row, links):
        self.setup_sheet()
        img0 = Image("__temp0.png")
        img0.width, img0.height = 132, 132
        img1 = Image("__temp221.png")
        img1.width, img1.height = 132, 132
        self.ws.add_image(img0, f"I{row}")
        self.ws.add_image(img1, f"J{row}")
        self.ws[f"K{row}"].hyperlink = links[0]
        self.ws[f"L{row}"].hyperlink = links[1]
        self.ws[f'K{row}'].value = links[0]
        self.ws[f'L{row}'].value = links[1]
        self.ws[f'K{row}'].style = "Hyperlink"
        self.ws[f'L{row}'].style = "Hyperlink"
        """
        self.ws[f'M{row}'].value = prices[0]
        self.ws[f'N{row}'].value = prices[1]
        """
        self.doc.save(self.tmp)
        MW.count += 1

    def final_save(self):
        if not os.path.isdir(os.getcwd() + "\\result\\"):
            os.makedirs(os.getcwd() + "\\result")
        shutil.copy(self.tmp, self.res)
        # os.remove(self.tmp)

    def __init__(self):
        fd = MW.files_dir[:-1]
        fn = fd[fd.rindex('/') + 1:]
        # od = fd[:fd.rindex('/')][:fd[:fd.rindex('/')].rindex('/')] + "\\result\\"
        self.set_directory(fn)

    def set_directory(self, filename):
        self.filename = filename
        self.tmp = os.getcwd() + f"\\_temp_{filename}_result.xlsx"
        self.res = os.getcwd() + "\\result\\" + self.filename + "_result.xlsx"

    def create_template(self):
        self.doc = openpyxl.load_workbook(MW.excelFilepath)
        self.ws = self.doc.active
        self.ws['I1'] = "사진1"
        self.ws['J1'] = "사진2"
        self.ws['K1'] = "링크1"
        self.ws['L1'] = "링크2"
        self.ws['M1'] = "구매수/가격1"
        self.ws['N1'] = "구매수/가격2"
        self.ws['O1'] = "타오바오 상품명"
        self.ws.column_dimensions['I'].width = 17.63
        self.ws.column_dimensions['J'].width = 17.63
        self.doc.save(self.tmp)

class MainWindow(QMainWindow, Ui_Window):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.onlyInt = QIntValidator()
        self.files_dir = ''
        self._files_dir = ''
        self.excelFilepath = ''
        self.btn_file.clicked.connect(self.select_dir)
        self.skipPush.stateChanged.connect(self.fix_state)
        self.skipRename.clicked.connect(self.fix_state_)
        self.multiFiles.stateChanged.connect(self.multi_file_method)
        self.btn_start.setEnabled(False)
        self.multi_files_report = False
        self.le_indexmax.setValidator(self.onlyInt)
        self.le_indexmin.setValidator(self.onlyInt)
        self.le_ocv_offset.setValidator(self.onlyInt)
        self.loadDelay: int = 12
        self.fetchDelay: int = 0  # TODO: indexRange
        self.indexRange: tuple = (0, 0)
        self.opencv_offset: int = 0
        self.opencv_confidence: float = 0.9

        self.count = 0

    def multi_file_method(self):
        if self.multiFiles.isChecked():
            self.multi_files_report = True
            self.lb_multifile.setHidden(False)
            if self.files_dir and self.files_dir != '/':
                self.btn_start.setEnabled(True)
            self.btn_start.clicked.connect(self.multi_start)
            self.btn_index.setEnabled(False)
            self.skipRename.setChecked(False)
            self.skipPush.setChecked(False)
            self.skipPush.setEnabled(False)
            self.skipRename.setEnabled(False)
        else:
            self.lb_multifile.setHidden(True)
            if not self.multiFiles.isChecked():
                self.btn_start.setEnabled(False)
            self.btn_start.clicked.connect(self.start)
            self.btn_index.clicked.connect(self.index)
            self.btn_index.setEnabled(True)
            self.skipPush.setEnabled(True)
            self.skipRename.setEnabled(True)

    def start(self):
        self.btn_start.setEnabled(False)
        if not self.index_validator():
            self.btn_start.setEnabled(True)
            self.warning_error01()
            return
        if not self.files_dir:
            self.btn_start.setEnabled(True)
            self.warning_error02()
            return
        self.loadDelay = int(self.le_loaddelay.text())
        self.fetchDelay = int(self.le_fetchdelay.text())
        self.error_report_number = str(self.le_erhp.text())
        self.opencv_offset = int(self.le_ocv_offset.text())
        self.opencv_confidence = float(self.le_ocv_confidence.text())
        self.worker = Worker(self.opencv_offset, self.opencv_confidence)
        self.worker.start()

    def select_dir(self):
        self.original_files_dir = QFileDialog.getExistingDirectory(
            parent=self,
            caption="Select a folder"
        ) + '/'
        self.files_dir = os.getcwd() + "/temp/data/"
        shutil.copytree_noattr(src=self.original_files_dir, dst=self.files_dir)
        self.output("Copied FileTree to /temp/data/")
        self.lb_filedir.setText(self.original_files_dir)
        self.lb_filedir.repaint()
        self.files_dir_ = self.files_dir
        if self.multiFiles.isChecked():
            self.btn_start.setEnabled(True)

    def multi_start(self):
        self.worker_finished = False
        self.multi_worker = MultiWorker()
        self.multi_worker.start()

    def index_validator(self):
        i_max = int(self.le_indexmax.text())
        # iMin = int(self.le_indexmin.text())
        if i_max == 0:
            return True
        else:
            """
            if iMin < 0 or iMin > i_max:
                return False
            """
            return False
        # return True

    def fix_state(self):
        if self.skipPush.isChecked():
            self.skipRename.setChecked(True)

    def fix_state_(self):
        if self.skipPush.isChecked() and not self.skipRename.isChecked():
            self.skipRename.setChecked(False)
            self.skipPush.setChecked(False)

    def index(self):
        errors = []
        try:
            self.ex = Excel()
            files = os.listdir(self.files_dir)
            self.real_images = []
            if self.skipPush.isChecked() and self.skipRename.isChecked():
                for file in files:
                    if ".xlsx" in file:
                        if "~$" not in file[0:3]:
                            self.excelFilepath = self.files_dir + file
                self.ex.create_template()
                self.ex.setup_sheet()
                for cellObj in reversed(list(MW.ex.ws.columns)[0]):
                    try:
                        index = cellObj.value
                        if (index and index != "Index") or index == 0:
                            self.real_images.append(int(index))
                    except ValueError:
                        self.output("ERROR: Invalid Literal!")
                        return
            elif self.skipRename.isChecked():
                os.system(Rf"platform-tools\adb.exe shell mkdir sdcard/taobao/{self.ex.filename}/")
                for file in files:
                    if ".xlsx" in file:
                        if "~$" not in file[0:3]:
                            self.excelFilepath = self.files_dir + file
                self.ex.create_template()
                self.ex.setup_sheet()
                for cellObj in reversed(list(MW.ex.ws.columns)[0]):
                    try:
                        index = cellObj.value
                        if (index and index != "Index") or index == 0:
                            self.real_images.append(int(index))
                    except ValueError:
                        self.output("ERROR: Invalid Literal!")
                        return
                for index in self.real_images:
                    src_filename = self.files_dir + f"{int(index):04}.jpg"
                    dst_filename = self.files_dir + f"_{int(index):04}.jpg"
                    manipulate_exif.set_time(index, src_filename, dst_filename)
                    os.remove(src_filename)
                    os.system(R"platform-tools\adb.exe push " + dst_filename +
                              f" sdcard/taobao/{self.ex.filename}/{int(index):04}.jpg")
            else:
                os.system(Rf"platform-tools\adb.exe shell mkdir sdcard/taobao/{self.ex.filename}/")
                for file in files:
                    if ".xlsx" not in file:
                        index = file[file.rindex("_") + 1:file.rindex(".jpg")]
                        os.rename(self.files_dir + file, self.files_dir + f"{int(index):04}.jpg")
                    else:
                        if "~$" not in file[0:3]:
                            self.excelFilepath = self.files_dir + file
                self.ex.create_template()
                self.ex.setup_sheet()
                for cellObj in reversed(list(MW.ex.ws.columns)[0]):
                    try:
                        index = cellObj.value
                        if (index and index != "Index") or index == 0:
                            self.real_images.append(int(index))
                    except ValueError:
                        self.output("ERROR: Invalid Literal!")
                        return
                for index in self.real_images:
                    src_filename = self.files_dir + f"{int(index):04}.jpg"
                    dst_filename = self.files_dir + f"_{int(index):04}.jpg"
                    manipulate_exif.set_time(index, src_filename, dst_filename)
                    os.remove(src_filename)
                    os.system(R"platform-tools\adb.exe push " + dst_filename +
                              f" sdcard/taobao/{self.ex.filename}/{int(index):04}.jpg")
            self.progressBar.setValue(0)
            self.progressBar.setRange(0, len(self.real_images))
            self.btn_start.setEnabled(True)
            self.output(f"Index Done; found {len(self.real_images)} items")
            if self.skipPush.isChecked():
                self.output("Skipped Push")
            else:
                self.output(f"Pushed {len(self.real_images)} items")
            self.btn_index.setEnabled(False)
        except Error as err:
            errors.extend(err.args[0])
        except OSError as why:
            errors.append(str(why))
        if errors:
            shutil.rmtree("temp/")
            raise Error(errors)

    def output(self, msg, error=False):
        if error:
            log.error(msg)
        else:
            log.info(msg)
        self.textBrowser.append(f"[{str(datetime.datetime.now())[11:19]}] " + msg)
        self.textBrowser.moveCursor(QTextCursor.End)

    def warning_error01(self):
        QMessageBox.warning(self, "[E01] FunctionError", "Custom Index Range is not Supported")
        # QMessageBox.warning(self, "[E01] IOError", "Check Index Range!")

    def warning_error02(self):
        QMessageBox.warning(self, "[E02] IOError", "No File Directory Selected!")

class Worker(QThread):
    def __init__(self, offset, confidence):
        super().__init__()
        log.info("Worker Thread Initiated")
        self.itemCoordinates = []
        self.itemIndices = []
        self.count = 0
        self.consecutiveSkips = 0
        self.skips = []
        self.offset = offset if offset <= 20 else 20
        self.confidence = confidence

    def run(self):
        self.pre_main()
        done = []
        skip_constant = 0
        for (count, index) in enumerate(MW.real_images):
            if (count + skip_constant) % 8 == 0 and count != 0:
                log.info("Restarting Debugger; count=%s, skip_constant=%s", count, skip_constant)
                if done:
                    for i in done:
                        os.system(R"platform-tools\adb.exe shell rm "
                                  f"sdcard/taobao/{MW.ex.filename}/{i:04}.jpg")
                    MW.output(f"Deleted Images: ~{done[len(done) - 1]:04}.jpg")
                else:
                    MW.output("Deleted Nothing")
                os.system(R"platform-tools\adb.exe devices")
                os.system(R"platform-tools\adb.exe shell am force-stop com.taobao.taobao")
                MW.output("Restarted Debugger")
                time.sleep(2)
                done = []
                self.launchTaobaoApp()
                time.sleep(3)
                Macro.single_click(Macro.btn_camera)
                time.sleep(3)
                old_gallery = pyautogui.locateOnScreen(
                    BASE_DIR + "taobaoOldGallery.png",
                    region=(
                        140 - self.offset,
                        700 - self.offset,
                        160 + self.offset,
                        110 + self.offset
                    ),
                    confidence=self.confidence)
                Macro.single_click(Macro.btn_gallery if old_gallery else Macro.btn_gallery_lower)
                time.sleep(3)

            done.append(index)
            Macro.click(self.itemCoordinates[(count + skip_constant) % 8])
            time.sleep(MW.loadDelay)
            Macro.before_locate()
            if not pyautogui.locateOnScreen(
                    BASE_DIR + "taobaoItemList.png",
                    region=(0, 30 - self.offset, 460 + self.offset, 400 + self.offset),
                    confidence=self.confidence):
                MW.output("Sleeping Extra Time")
                time.sleep(20 - MW.loadDelay)
            if not self.fetch_data(index, len(MW.real_images) - count + 1):
                skip_constant = 7 - (count % 8)
                self.skips.append((index, len(MW.real_images) - count + 1))
                if self.consecutiveSkips == 3:
                    self.send_report("error", done_count=count)
                    return
                self.consecutiveSkips += 1
            else:
                self.consecutiveSkips = 0
            tries = 0
            while True:
                if tries > 6:
                    return
                Macro.single_click(Macro.btn_escape)
                time.sleep(3)
                Macro.before_locate()
                if pyautogui.locateOnScreen(
                        BASE_DIR + "taobaoCamera.png",
                        confidence=self.confidence,
                        region=(
                                300 - self.offset,
                                60 - self.offset,
                                150 + self.offset,
                                100 + self.offset
                        )):
                    break
                if pyautogui.locateOnScreen(
                        BASE_DIR + "taobaoMain.png",
                        confidence=self.confidence,
                        region=(
                                270 - self.offset,
                                850 - self.offset,
                                170 + self.offset,
                                80 + self.offset
                        )):
                    MW.output("Returned from Main")
                    Macro.click(Macro.btn_camera)
                    time.sleep(2)
                    break
                tries += 1
            old_gallery = pyautogui.locateOnScreen(
                BASE_DIR + "taobaoOldGallery.png",
                region=(
                    140 - self.offset,
                    700 - self.offset,
                    160 + self.offset,
                    110 + self.offset
                ),
                confidence=self.confidence)
            Macro.single_click(Macro.btn_gallery if old_gallery else Macro.btn_gallery_lower)
            time.sleep(MW.fetchDelay)
            self.last_done = count

        self.end_time = datetime.datetime.now()
        self.total_time = self.end_time - self.start_time
        MW.output(f"Fetch Done; total={MW.count}; "
                  "{self.total_time.total_seconds() // 3600}h "
                  "{self.total_time.total_seconds() // 60}m")
        for i in done:
            os.system(R"platform-tools\adb.exe shell rm "
                      f"sdcard/taobao/{MW.ex.filename}/{i:04}.jpg")
        MW.output(f"Deleted Images: ~{done[len(done) - 1]:04}.jpg")
        MW.ex.final_save()
        if not MW.multi_files_report:
            self.send_report("success", done_count=self.last_done)
        else:
            if MW.multi_worker.cnt == MW.multi_worker.total_files:
                self.send_report("success", done_count=self.last_done)
        MW.output("Killing ADB")
        os.system(R"platform-tools\adb.exe kill-server")
        MW.output("Deleting SubThread")
        del self.debugger
        MW.worker_finished = True
        self.deleteLater()
        self.quit()

    def fetch_data(self, index, write_row) -> 0 or 1:
        links = []
        # urls = []
        Macro.before_locate()
        error = pyautogui.locateOnScreen(
            BASE_DIR + "taobaoError.png",
            region=(
                20 - self.offset,
                400 - self.offset,
                410 + self.offset,
                140 + self.offset
            ),
            confidence=self.confidence)
        if error:
            MW.output("Error Popup Detected", error=True)
            Macro.click((360, 500))
            time.sleep(15)
            if pyautogui.locateOnScreen(
                    BASE_DIR + "taobaoErrorLoading.png",
                    region=(
                            60 - self.offset,
                            300 - self.offset,
                            300 + self.offset,
                            450 + self.offset
                    ),
                    confidence=self.confidence - 0.05):
                MW.output("errorLoading Detected", error=True)
                time.sleep(30)
        Macro.before_locate()
        error1 = pyautogui.locateOnScreen(
            BASE_DIR + "taobaoError_.png",
            region=(
                20 - self.offset,
                400 - self.offset,
                410 + self.offset,
                140 + self.offset
            ),
            confidence=self.confidence)
        if error1:
            MW.output("Error Popup Detected", error=True)
            Macro.click((360, 500))
            time.sleep(15)
            if pyautogui.locateOnScreen(
                    BASE_DIR + "taobaoErrorLoading.png",
                    region=(
                            60 - self.offset,
                            300 - self.offset,
                            300 + self.offset,
                            450 + self.offset
                    ),
                    confidence=self.confidence - 0.05):
                MW.output("errorLoading Detected", error=True)
                time.sleep(30)
        offset = 0
        Macro.before_locate()
        if not pyautogui.locateOnScreen(
                BASE_DIR + "taobaoCharacterResult.png",
                region=(
                        0,
                        240 - self.offset,
                        100 + self.offset,
                        60 + self.offset
                ),
                confidence=self.confidence):
            if not pyautogui.locateOnScreen(
                    BASE_DIR + "taobaoFilter1.png",
                    region=(
                            0,
                            170 - self.offset,
                            330 + self.offset,
                            60 + self.offset
                    ),
                    confidence=self.confidence - 0.05):
                if not pyautogui.locateOnScreen(
                        BASE_DIR + "taobaoFilter2.png",
                        region=(
                                0,
                                170 - self.offset,
                                330 + self.offset,
                                60 + self.offset
                        ),
                        confidence=self.confidence - 0.05):
                    if not pyautogui.locateOnScreen(
                            BASE_DIR + "taobaoFilter3.png",
                            region=(
                                    0,
                                    170 - self.offset,
                                    330 + self.offset,
                                    60 + self.offset
                            ),
                            confidence=self.confidence - 0.05):
                        if not pyautogui.locateOnScreen(
                                BASE_DIR + "taobaoFilter4.png",
                                region=(
                                        0,
                                        170 - self.offset,
                                        330 + self.offset,
                                        60 + self.offset
                                ),
                                confidence=self.confidence - 0.05):
                            if not pyautogui.locateOnScreen(
                                    BASE_DIR + "taobaoFilter5.png",
                                    region=(
                                            0,
                                            170 - self.offset,
                                            330 + self.offset,
                                            60 + self.offset
                                    ),
                                    confidence=self.confidence - 0.05):
                                if not pyautogui.locateOnScreen(
                                        BASE_DIR + "taobaoFilter6.png",
                                        region=(
                                                0,
                                                170 - self.offset,
                                                330 + self.offset,
                                                60 + self.offset
                                        ),
                                        confidence=self.confidence - 0.05):
                                    if not pyautogui.locateOnScreen(
                                            BASE_DIR + "taobaoItemList.png",
                                            region=(
                                                    0,
                                                    30 - self.offset,
                                                    460 + self.offset,
                                                    200 + self.offset
                                            ),
                                            confidence=self.confidence):
                                        MW.output("Offset Level = 2")
                                        offset = 100
                                    else:
                                        MW.output("Offset Level = 1")
                                        offset = 50
        else:
            MW.output("Offset Level = -1")
            offset = -170
        for result in (Macro.result_1, Macro.result_2):
            c = 221 if result[2] == 2 else 0
            Macro.before_locate()
            time.sleep(1)  # Load Images
            pyautogui.screenshot(f"__temp{c}.png",
                                 region=(
                                     1 + c,
                                     228 - offset,
                                     216,
                                     216
                                 ))
            if offset == -170:
                Macro.click((result[0], result[1] + 200))
            else:
                Macro.click(result)
            time.sleep(2.5)
            if pyautogui.locateOnScreen(
                    BASE_DIR + "taobaoVerify.png",
                    region=(
                            30 - self.offset,
                            200 - self.offset,
                            400 + self.offset,
                            600 + self.offset
                    ),
                    confidence=self.confidence):
                Macro.single_click(Macro.btn_escape)
                time.sleep(2)
            Macro.click(Macro.btn_share)
            time.sleep(2)
            try:
                copy_link_btn = pyautogui.center(pyautogui.locateOnScreen(
                    BASE_DIR + "taobaoCopyLink.png",
                    confidence=self.confidence
                ))
            except TypeError:
                MW.count += 1
                ts = datetime.datetime.now().timestamp()
                if not os.path.isdir(os.getcwd() + "\\screenshots\\"):
                    os.makedirs(os.getcwd() + "\\screenshots")
                pyautogui.screenshot(Rf"screenshots\{ts}.png")
                log.error("Saved Screenshot: %s.png", ts)
                MW.output(f"Skip item[{index}]; row={write_row}, count={MW.count}", error=True)
                return
            Macro.click(copy_link_btn)
            time.sleep(2)
            for _ in range(0, 2):
                Macro.single_click(Macro.btn_escapeItemDetail)
                time.sleep(1.5)
            tries = 0
            while True:
                if tries > 6:
                    return
                if not pyautogui.locateOnScreen(
                        BASE_DIR + "taobaoItemDetail.png",
                        region=(0, 830 - self.offset, 180 + self.offset, 80 + self.offset),
                        confidence=self.confidence):
                    if not pyautogui.locateOnScreen(
                            BASE_DIR + "taobaoItemDetail2.png",
                            region=(0, 845 - self.offset, 165 + self.offset, 60 + self.offset),
                            confidence=self.confidence):
                        break
                Macro.single_click(Macro.btn_escapeItemDetail)
                time.sleep(3)
                tries += 1
            link = pyperclip.paste()
            try:
                links.append(link[4:link.index('「') - 8])
            except ValueError:
                links.append(link)
        # prices = []
        """
        for link in links:
            url, price = self.redirect_link(link)
            urls.append(url)
            prices.append(price)
        """
        MW.ex.write(write_row, links)
        os.remove("__temp0.png")
        os.remove("__temp221.png")
        MW.progressBar.setValue(MW.count)
        MW.output(f"Item[{index}]; row={write_row}, count={MW.count}")
        return 1

    def pre_main(self):
        self.connect_debugger()
        MW.output("Attached Debugger to the device")
        time.sleep(5)
        self.launchTaobaoApp()
        MW.output("Launched Taobao App")
        Macro().move_window()
        time.sleep(3)
        Macro.click(Macro.btn_camera)
        time.sleep(3)
        old_gallery = pyautogui.locateOnScreen(
            BASE_DIR + "taobaoOldGallery.png",
            region=(
                140 - self.offset,
                700 - self.offset,
                160 + self.offset,
                110 + self.offset
            ),
            confidence=self.confidence
        )
        Macro.single_click(Macro.btn_gallery if old_gallery else Macro.btn_gallery_lower)
        time.sleep(3)
        for i in range(0, 8):
            x = Macro.item_1[0] + ((i % 4) * 106)
            y = Macro.item_1[1] + (106 if (i // 4) % 2 == 1 else 0)
            self.itemCoordinates.append((x, y))
        MW.output("Start Fetch")
        self.start_time = datetime.datetime.now()

    """
    @staticmethod
    def redirect_link(link) -> tuple:
        r = requests.get(link)
        r.encoding = "utf-8"
        res = r.text
        res1 = res[res.index("var url = ") + 11:]
        price = res1[res1.index("&price")+7:res1.index("&source")]
        return res1[:res1.index("&price")], price
    """

    @staticmethod
    def launch_taobao_app():
        os.system(R"platform-tools\adb.exe shell am start -n "
                  "com.taobao.taobao/com.taobao.tao.TBMainActivity")

    @staticmethod
    def send_report(report_type, done_count=0):
        os.system(R"platform-tools\adb.exe shell am force-stop "
                  "com.samsung.android.messaging")
        time.sleep(3)
        os.system(R"platform-tools\adb.exe shell am start -n "
                  "com.samsung.android.messaging/com.android.mms.ui.ConversationComposer")
        ts = datetime.datetime.now()
        time.sleep(3)
        Macro.single_click((390, 790))
        time.sleep(2)
        pyautogui.typewrite(str(MW.error_report_number), interval=0.02)
        time.sleep(3)
        Macro.single_click((200, 500))
        if report_type == "error":
            pyautogui.typewrite(f"{ts}-!!!!!!400!!!!!!({done_count})", interval=0.01)
        elif report_type == "success":
            pyautogui.typewrite(f"{ts}-______200______({done_count})", interval=0.01)
        time.sleep(3)
        Macro.single_click((410, 510))

    def connect_debugger(self):
        self.debugger = Debugger()
        self.debugger.start()

class MultiWorker(QThread):
    def __init__(self):
        super().__init__()
        log.info("MultiWorker Thread Initiated")
        self.cnt = 1

    def run(self):
        multi_files = os.listdir(MW.files_dir_)
        self.total_files = len(multi_files)
        MW.output(f"MultiFetch Start; fileCount={self.total_files}")
        for (cnt, file) in enumerate(multi_files, start=1):
            self.cnt = cnt
            MW.lb_multifile.setText(f"File {cnt}/{self.total_files}")
            MW.files_dir = MW.files_dir_ + file + '\\'
            MW.index()
            MW.start()
            while True:
                time.sleep(10)
                if MW.worker_finished:
                    MW.output(f"File {cnt}/{len(multi_files)} Done")
                    break
            MW.worker_finished = False
            del MW.worker
            time.sleep(5)
            MW.count = 0
        shutil.rmtree("temp/")

class Debugger(threading.Thread):
    def __init__(self):
        super().__init__()

    def run(self):
        os.system(R"platform-tools\scrcpy.exe -d")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    MW = MainWindow()
    log = Logger().logger
    MW.show()
    app.exec_()
